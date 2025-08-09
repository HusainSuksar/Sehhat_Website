import os
import tempfile
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.views.decorators.http import require_http_methods
from django.utils.decorators import method_decorator
from django.views.generic import ListView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Q
from django.conf import settings
import json

from .models import BulkUploadSession, BulkUploadRecord, UploadTemplate
from .services import BulkUploadService, FileProcessor
from accounts.models import User


class AdminRequiredMixin(UserPassesTestMixin):
    """Mixin to ensure user is admin or staff"""
    
    def test_func(self):
        return self.request.user.is_authenticated and (
            self.request.user.is_admin or 
            self.request.user.is_superuser or
            self.request.user.role == 'badri_mahal_admin'
        )
    
    def handle_no_permission(self):
        messages.error(self.request, "You don't have permission to access bulk upload functionality.")
        return redirect('accounts:dashboard')


@method_decorator(login_required, name='dispatch')
class BulkUploadListView(AdminRequiredMixin, ListView):
    """List all bulk upload sessions"""
    model = BulkUploadSession
    template_name = 'bulk_upload/upload_list.html'
    context_object_name = 'uploads'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = BulkUploadSession.objects.all().select_related('uploaded_by')
        
        # Filter by upload type
        upload_type = self.request.GET.get('type')
        if upload_type:
            queryset = queryset.filter(upload_type=upload_type)
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Search by filename or user
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(original_filename__icontains=search) |
                Q(uploaded_by__first_name__icontains=search) |
                Q(uploaded_by__last_name__icontains=search)
            )
        
        return queryset.order_by('-started_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['upload_types'] = BulkUploadSession.UPLOAD_TYPE_CHOICES
        context['status_choices'] = BulkUploadSession.STATUS_CHOICES
        context['current_filters'] = {
            'type': self.request.GET.get('type', ''),
            'status': self.request.GET.get('status', ''),
            'search': self.request.GET.get('search', ''),
        }
        return context


@method_decorator(login_required, name='dispatch')
class BulkUploadDetailView(AdminRequiredMixin, DetailView):
    """View details of a specific bulk upload session"""
    model = BulkUploadSession
    template_name = 'bulk_upload/upload_detail.html'
    context_object_name = 'upload'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get records with pagination
        records = self.object.records.all()
        
        # Filter by status
        status_filter = self.request.GET.get('record_status')
        if status_filter:
            records = records.filter(status=status_filter)
        
        context['records'] = records[:100]  # Limit for performance
        context['record_status_choices'] = BulkUploadRecord.STATUS_CHOICES
        context['current_record_filter'] = status_filter or ''
        
        # Statistics
        context['stats'] = {
            'success_rate': self.object.get_success_rate(),
            'total_records': records.count(),
            'failed_records': records.filter(status='failed'),
            'success_records': records.filter(status='success'),
        }
        
        return context


@login_required
def bulk_upload_create(request):
    """Create a new bulk upload"""
    if not (request.user.is_admin or request.user.role == 'badri_mahal_admin'):
        messages.error(request, "You don't have permission to create bulk uploads.")
        return redirect('accounts:dashboard')
    
    if request.method == 'POST':
        upload_type = request.POST.get('upload_type')
        uploaded_file = request.FILES.get('file')
        
        if not upload_type or not uploaded_file:
            messages.error(request, "Please select an upload type and file.")
            return render(request, 'bulk_upload/upload_create.html', get_upload_context())
        
        # Validate file type
        allowed_extensions = ['xlsx', 'xls', 'csv']
        file_extension = uploaded_file.name.split('.')[-1].lower()
        if file_extension not in allowed_extensions:
            messages.error(request, f"File type '{file_extension}' not supported. Please use: {', '.join(allowed_extensions)}")
            return render(request, 'bulk_upload/upload_create.html', get_upload_context())
        
        # Save file temporarily
        try:
            file_content = uploaded_file.read()
            
            # Create a proper temporary file path
            import tempfile
            
            # Create temporary file
            temp_fd, temp_file_path = tempfile.mkstemp(
                suffix=f'.{file_extension}',
                prefix=f'bulk_upload_{request.user.id}_'
            )
            
            # Write file content to temporary file
            with os.fdopen(temp_fd, 'wb') as temp_file:
                temp_file.write(file_content)
            
            # Create upload session
            session = BulkUploadService.create_upload_session(
                user=request.user,
                upload_type=upload_type,
                file_path=temp_file_path,
                filename=uploaded_file.name,
                file_size=len(file_content)
            )
            
            messages.success(request, f"File uploaded successfully. Upload session #{session.id} created.")
            return redirect('bulk_upload:detail', pk=session.id)
            
        except Exception as e:
            messages.error(request, f"Error uploading file: {str(e)}")
            return render(request, 'bulk_upload/upload_create.html', get_upload_context())
    
    return render(request, 'bulk_upload/upload_create.html', get_upload_context())


@login_required
@require_http_methods(["POST"])
def bulk_upload_process(request, pk):
    """Start processing a bulk upload session"""
    if not (request.user.is_admin or request.user.role == 'badri_mahal_admin'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    session = get_object_or_404(BulkUploadSession, pk=pk)
    
    if session.status != 'pending':
        return JsonResponse({'error': 'Upload session is not in pending status'}, status=400)
    
    try:
        # Process the upload in background (in a real app, use Celery)
        BulkUploadService.process_upload(session)
        
        return JsonResponse({
            'success': True,
            'message': 'Upload processed successfully',
            'stats': {
                'total': session.total_rows,
                'successful': session.successful_rows,
                'failed': session.failed_rows,
                'skipped': session.skipped_rows,
                'success_rate': session.get_success_rate()
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Processing failed: {str(e)}'}, status=500)


@login_required
def bulk_upload_preview(request, pk):
    """Preview file contents before processing"""
    if not (request.user.is_admin or request.user.role == 'badri_mahal_admin'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    session = get_object_or_404(BulkUploadSession, pk=pk)
    
    try:
        # Check if file exists
        if not session.file_path or not os.path.exists(session.file_path):
            return JsonResponse({'error': 'File not found. Please upload again.'}, status=404)
        
        # Read first few rows for preview
        file_processor = FileProcessor(session.file_path, session.original_filename.split('.')[-1])
        data = file_processor.read_file()
        
        # Limit preview to first 10 rows
        preview_data = data[:10]
        
        # Get headers
        headers = list(preview_data[0].keys()) if preview_data else []
        if '_row_number' in headers:
            headers.remove('_row_number')
        
        # Validate headers against template (if exists)
        valid_headers, missing_headers = BulkUploadService.validate_file_headers(session.upload_type, headers)
        
        return JsonResponse({
            'success': True,
            'headers': headers,
            'preview_data': preview_data,
            'total_rows': len(data),
            'valid_headers': valid_headers,
            'missing_headers': missing_headers
        })
        
    except FileNotFoundError:
        return JsonResponse({'error': 'File not found. Please upload again.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Failed to preview file: {str(e)}'}, status=500)


@login_required
def download_template(request, upload_type):
    """Download Excel template for specific upload type"""
    if not (request.user.is_admin or request.user.role == 'badri_mahal_admin'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        template = UploadTemplate.objects.get(upload_type=upload_type, is_active=True)
        
        # Create Excel file with template
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{template.name} Template"
        
        # Add headers
        headers = template.get_all_columns()
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Add sample data if available
        if template.sample_data:
            for row_num, sample_row in enumerate(template.sample_data[:5], 2):
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col_num, value=sample_row.get(header, ''))
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{upload_type}_template.xlsx"'
        wb.save(response)
        return response
        
    except UploadTemplate.DoesNotExist:
        raise Http404("Template not found")
    except Exception as e:
        return JsonResponse({'error': f'Failed to generate template: {str(e)}'}, status=500)


@login_required
def bulk_upload_delete(request, pk):
    """Delete a bulk upload session"""
    if not (request.user.is_admin or request.user.role == 'badri_mahal_admin'):
        messages.error(request, "Permission denied.")
        return redirect('bulk_upload:list')
    
    session = get_object_or_404(BulkUploadSession, pk=pk)
    
    if request.method == 'POST':
        # Delete associated file
        if session.file_path and os.path.exists(session.file_path):
            try:
                os.unlink(session.file_path)
            except OSError:
                pass  # File already deleted or doesn't exist
        
        session.delete()
        messages.success(request, "Upload session deleted successfully.")
        return redirect('bulk_upload:list')
    
    return render(request, 'bulk_upload/upload_confirm_delete.html', {'upload': session})


def get_upload_context():
    """Get context data for upload forms"""
    return {
        'upload_types': BulkUploadSession.UPLOAD_TYPE_CHOICES,
        'templates': BulkUploadService.get_upload_templates(),
        'max_file_size': getattr(settings, 'FILE_UPLOAD_MAX_MEMORY_SIZE', 10 * 1024 * 1024),  # 10MB default
    }


@login_required
def bulk_upload_api_status(request, pk):
    """Get real-time status of upload processing"""
    if not (request.user.is_admin or request.user.role == 'badri_mahal_admin'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    session = get_object_or_404(BulkUploadSession, pk=pk)
    
    return JsonResponse({
        'status': session.status,
        'total_rows': session.total_rows,
        'successful_rows': session.successful_rows,
        'failed_rows': session.failed_rows,
        'skipped_rows': session.skipped_rows,
        'success_rate': session.get_success_rate(),
        'completed_at': session.completed_at.isoformat() if session.completed_at else None,
        'processing_log': session.processing_log[-10:] if session.processing_log else []  # Last 10 entries
    })
