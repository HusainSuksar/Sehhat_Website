"""
Bulk Upload Services for processing Excel/CSV files
"""
import csv
import json
from typing import Dict, List, Any, Tuple, Optional
from datetime import datetime, date
from decimal import Decimal
import re

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone

from .models import BulkUploadSession, BulkUploadRecord, UploadTemplate
from accounts.models import User
from accounts.services import MockITSService
from students.models import Student
from moze.models import Moze
from doctordirectory.models import Doctor
from mahalshifa.models import Patient, MedicalRecord


class FileProcessor:
    """Base class for processing uploaded files"""
    
    def __init__(self, file_path: str, file_type: str):
        self.file_path = file_path
        self.file_type = file_type.lower()
    
    def read_file(self) -> List[Dict[str, Any]]:
        """Read file and return list of dictionaries"""
        if self.file_type in ['xlsx', 'xls']:
            return self._read_excel()
        elif self.file_type == 'csv':
            return self._read_csv()
        else:
            raise ValueError(f"Unsupported file type: {self.file_type}")
    
    def _read_excel(self) -> List[Dict[str, Any]]:
        """Read Excel file"""
        data = []
        
        if self.file_type == 'xlsx':
            try:
                import openpyxl  # Lazy import to avoid hard dependency at module import time
            except ImportError as exc:
                raise ImportError("openpyxl is required to process .xlsx files. Please install openpyxl.") from exc
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).strip() if cell.value else "")
            
            # Read data rows
            for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                row_data = {}
                for col_num, cell in enumerate(row):
                    if col_num < len(headers) and headers[col_num]:
                        value = cell.value
                        if isinstance(value, datetime):
                            value = value.date().isoformat()
                        elif isinstance(value, date):
                            value = value.isoformat()
                        row_data[headers[col_num]] = str(value).strip() if value is not None else ""
                
                if any(row_data.values()):  # Skip empty rows
                    row_data['_row_number'] = row_num
                    data.append(row_data)
        
        elif self.file_type == 'xls':
            try:
                import xlrd  # Lazy import
            except ImportError as exc:
                raise ImportError("xlrd is required to process .xls files. Please install xlrd.") from exc
            workbook = xlrd.open_workbook(self.file_path)
            sheet = workbook.sheet_by_index(0)
            
            # Get headers from first row
            headers = []
            for col in range(sheet.ncols):
                headers.append(str(sheet.cell(0, col).value).strip())
            
            # Read data rows
            for row_num in range(1, sheet.nrows):
                row_data = {}
                for col_num in range(sheet.ncols):
                    if col_num < len(headers) and headers[col_num]:
                        cell = sheet.cell(row_num, col_num)
                        value = cell.value
                        
                        # Handle date values
                        if cell.ctype == 3:  # Date
                            date_tuple = xlrd.xldate_as_tuple(value, workbook.datemode)
                            value = date(*date_tuple[:3]).isoformat()
                        
                        row_data[headers[col_num]] = str(value).strip() if value else ""
                
                if any(row_data.values()):  # Skip empty rows
                    row_data['_row_number'] = row_num + 1
                    data.append(row_data)
        
        return data
    
    def _read_csv(self) -> List[Dict[str, Any]]:
        """Read CSV file"""
        data = []
        
        with open(self.file_path, 'r', encoding='utf-8-sig') as file:
            # Try to detect delimiter
            sample = file.read(1024)
            file.seek(0)
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample).delimiter
            
            reader = csv.DictReader(file, delimiter=delimiter)
            
            for row_num, row in enumerate(reader, start=2):
                # Clean up the row data
                cleaned_row = {}
                for key, value in row.items():
                    clean_key = str(key).strip() if key else ""
                    clean_value = str(value).strip() if value else ""
                    if clean_key:
                        cleaned_row[clean_key] = clean_value
                
                if any(cleaned_row.values()):  # Skip empty rows
                    cleaned_row['_row_number'] = row_num
                    data.append(cleaned_row)
        
        return data


class DataProcessor:
    """Process and validate data for different entity types"""
    
    def __init__(self, upload_session: BulkUploadSession):
        self.session = upload_session
        self.upload_type = upload_session.upload_type
    
    def process_data(self, data: List[Dict[str, Any]]) -> None:
        """Process all data rows"""
        self.session.total_rows = len(data)
        self.session.status = 'processing'
        self.session.save()
        
        for row_data in data:
            row_number = row_data.pop('_row_number', 0)
            
            # Create record for tracking
            record = BulkUploadRecord.objects.create(
                session=self.session,
                row_number=row_number,
                raw_data=row_data
            )
            
            try:
                # Process based on upload type
                created_object = self._process_single_row(row_data, record)
                record.mark_success(created_object)
                
            except Exception as e:
                record.mark_failed(str(e))
                self.session.add_log_entry('error', f"Row {row_number}: {str(e)}", row_number)
        
        self.session.mark_completed()
    
    def _process_single_row(self, row_data: Dict[str, Any], record: BulkUploadRecord) -> Any:
        """Process a single row based on upload type"""
        if self.upload_type == 'users':
            return self._create_user(row_data, record)
        elif self.upload_type == 'students':
            return self._create_student(row_data, record)
        elif self.upload_type == 'doctors':
            return self._create_doctor(row_data, record)
        elif self.upload_type == 'patients':
            return self._create_patient(row_data, record)
        elif self.upload_type == 'moze':
            return self._create_moze(row_data, record)
        else:
            raise ValueError(f"Unsupported upload type: {self.upload_type}")
    
    def _create_user(self, row_data: Dict[str, Any], record: BulkUploadRecord) -> User:
        """Create a user from row data"""
        # Required fields
        required_fields = ['its_id', 'first_name', 'last_name', 'email', 'role']
        for field in required_fields:
            if not row_data.get(field):
                raise ValidationError(f"Missing required field: {field}")
        
        its_id = row_data['its_id'].strip()
        email = row_data['email'].strip().lower()
        
        # Check if user already exists
        if User.objects.filter(its_id=its_id).exists():
            raise ValidationError(f"User with ITS ID {its_id} already exists")
        
        if User.objects.filter(email=email).exists():
            raise ValidationError(f"User with email {email} already exists")
        
        # Validate role
        valid_roles = [choice[0] for choice in User.ROLE_CHOICES]
        if row_data['role'] not in valid_roles:
            raise ValidationError(f"Invalid role: {row_data['role']}")
        
        # Try to fetch additional data from ITS if available
        its_data = MockITSService.fetch_user_data(its_id)
        
        with transaction.atomic():
            user = User.objects.create_user(
                username=its_id,  # Use ITS ID as username
                email=email,
                first_name=row_data['first_name'].strip(),
                last_name=row_data['last_name'].strip(),
                its_id=its_id,
                role=row_data['role'],
                is_active=True,
                password=f"temp_{its_id}"  # Temporary password
            )
            
            # Add additional ITS data if available
            if its_data:
                for field, value in its_data.items():
                    if hasattr(user, field) and value:
                        setattr(user, field, value)
                user.save()
            
            # Add optional fields from row data
            optional_fields = ['mobile_number', 'occupation', 'qualification', 'idara', 'category']
            updated = False
            for field in optional_fields:
                if row_data.get(field) and hasattr(user, field):
                    setattr(user, field, row_data[field])
                    updated = True
            
            if updated:
                user.save()
        
        return user
    
    def _create_student(self, row_data: Dict[str, Any], record: BulkUploadRecord) -> Student:
        """Create a student from row data"""
        # Check if we have a user ITS ID or need to create user
        its_id = row_data.get('its_id', '').strip()
        if not its_id:
            raise ValidationError("ITS ID is required for student creation")
        
        # Try to get existing user or create one
        try:
            user = User.objects.get(its_id=its_id)
            if user.role != 'student':
                user.role = 'student'
                user.save()
        except User.DoesNotExist:
            # Create user first
            user_data = {
                'its_id': its_id,
                'first_name': row_data.get('first_name', f'Student{its_id}'),
                'last_name': row_data.get('last_name', ''),
                'email': row_data.get('email', f'{its_id}@student.example.com'),
                'role': 'student'
            }
            user = self._create_user(user_data, record)
        
        # Check if student profile already exists
        if hasattr(user, 'student'):
            raise ValidationError(f"Student profile for ITS ID {its_id} already exists")
        
        # Required fields for student
        student_id = row_data.get('student_id', its_id)
        academic_level = row_data.get('academic_level', 'undergraduate')
        enrollment_date = row_data.get('enrollment_date', timezone.now().date())
        
        # Parse enrollment date if it's a string
        if isinstance(enrollment_date, str):
            try:
                enrollment_date = datetime.strptime(enrollment_date, '%Y-%m-%d').date()
            except ValueError:
                enrollment_date = timezone.now().date()
        
        with transaction.atomic():
            student = Student.objects.create(
                user=user,
                student_id=student_id,
                academic_level=academic_level,
                enrollment_date=enrollment_date,
                enrollment_status=row_data.get('enrollment_status', 'active')
            )
            
            # Add expected graduation date if provided
            if row_data.get('expected_graduation'):
                try:
                    grad_date = datetime.strptime(row_data['expected_graduation'], '%Y-%m-%d').date()
                    student.expected_graduation = grad_date
                    student.save()
                except ValueError:
                    pass
        
        return student
    
    def _create_doctor(self, row_data: Dict[str, Any], record: BulkUploadRecord) -> Doctor:
        """Create a doctor from row data"""
        its_id = row_data.get('its_id', '').strip()
        if not its_id:
            raise ValidationError("ITS ID is required for doctor creation")
        
        # Try to get existing user or create one
        try:
            user = User.objects.get(its_id=its_id)
            if user.role != 'doctor':
                user.role = 'doctor'
                user.save()
        except User.DoesNotExist:
            user_data = {
                'its_id': its_id,
                'first_name': row_data.get('first_name', f'Dr{its_id}'),
                'last_name': row_data.get('last_name', ''),
                'email': row_data.get('email', f'{its_id}@doctor.example.com'),
                'role': 'doctor'
            }
            user = self._create_user(user_data, record)
        
        # Check if doctor profile already exists
        if Doctor.objects.filter(user=user).exists():
            raise ValidationError(f"Doctor profile for ITS ID {its_id} already exists")
        
        with transaction.atomic():
            doctor = Doctor.objects.create(
                user=user,
                specialization=row_data.get('specialization', 'General Medicine'),
                license_number=row_data.get('license_number', f'LIC{its_id}'),
                experience_years=int(row_data.get('experience_years', 0)),
                consultation_fee=Decimal(row_data.get('consultation_fee', '500.00')),
                is_available=row_data.get('is_available', 'true').lower() == 'true'
            )
            
            # Add optional fields
            optional_fields = ['qualifications', 'languages_spoken', 'clinic_address']
            for field in optional_fields:
                if row_data.get(field) and hasattr(doctor, field):
                    setattr(doctor, field, row_data[field])
            
            doctor.save()
        
        return doctor
    
    def _create_patient(self, row_data: Dict[str, Any], record: BulkUploadRecord) -> Patient:
        """Create a patient from row data"""
        # Required fields
        required_fields = ['its_id', 'first_name', 'last_name']
        for field in required_fields:
            if not row_data.get(field):
                raise ValidationError(f"Missing required field: {field}")
        
        its_id = row_data['its_id'].strip()
        
        # Check if patient already exists
        if Patient.objects.filter(its_id=its_id).exists():
            raise ValidationError(f"Patient with ITS ID {its_id} already exists")
        
        with transaction.atomic():
            patient = Patient.objects.create(
                its_id=its_id,
                first_name=row_data['first_name'].strip(),
                last_name=row_data['last_name'].strip(),
                date_of_birth=row_data.get('date_of_birth'),
                gender=row_data.get('gender', 'male'),
                phone_number=row_data.get('phone_number', ''),
                email=row_data.get('email', ''),
                address=row_data.get('address', ''),
                emergency_contact_name=row_data.get('emergency_contact_name', ''),
                emergency_contact_phone=row_data.get('emergency_contact_phone', '')
            )
        
        return patient
    
    def _create_moze(self, row_data: Dict[str, Any], record: BulkUploadRecord) -> Moze:
        """Create a Moze center from row data"""
        # Required fields
        required_fields = ['name', 'location']
        for field in required_fields:
            if not row_data.get(field):
                raise ValidationError(f"Missing required field: {field}")
        
        name = row_data['name'].strip()
        
        # Check if Moze already exists
        if Moze.objects.filter(name=name).exists():
            raise ValidationError(f"Moze with name '{name}' already exists")
        
        # Try to find aamil user
        aamil = None
        aamil_its_id = row_data.get('aamil_its_id')
        if aamil_its_id:
            try:
                aamil = User.objects.get(its_id=aamil_its_id, role='aamil')
            except User.DoesNotExist:
                raise ValidationError(f"Aamil with ITS ID {aamil_its_id} not found")
        else:
            # Use first available aamil
            aamil = User.objects.filter(role='aamil').first()
            if not aamil:
                raise ValidationError("No aamil users available")
        
        with transaction.atomic():
            moze = Moze.objects.create(
                name=name,
                location=row_data['location'].strip(),
                address=row_data.get('address', ''),
                aamil=aamil,
                capacity=int(row_data.get('capacity', 100)),
                contact_phone=row_data.get('contact_phone', ''),
                contact_email=row_data.get('contact_email', ''),
                is_active=row_data.get('is_active', 'true').lower() == 'true'
            )
            
            # Add coordinator if specified
            coordinator_its_id = row_data.get('coordinator_its_id')
            if coordinator_its_id:
                try:
                    coordinator = User.objects.get(its_id=coordinator_its_id, role='moze_coordinator')
                    moze.moze_coordinator = coordinator
                    moze.save()
                except User.DoesNotExist:
                    pass  # Skip if coordinator not found
        
        return moze


class BulkUploadService:
    """Main service for handling bulk uploads"""
    
    @staticmethod
    def create_upload_session(user, upload_type: str, file_path: str, filename: str, file_size: int, options: Dict = None) -> BulkUploadSession:
        """Create a new bulk upload session"""
        return BulkUploadSession.objects.create(
            upload_type=upload_type,
            uploaded_by=user,
            original_filename=filename,
            file_path=file_path,
            file_size=file_size,
            options=options or {}
        )
    
    @staticmethod
    def process_upload(session: BulkUploadSession) -> None:
        """Process an upload session"""
        try:
            # Read the file
            file_processor = FileProcessor(session.file_path, session.original_filename.split('.')[-1])
            data = file_processor.read_file()
            
            # Process the data
            data_processor = DataProcessor(session)
            data_processor.process_data(data)
            
        except Exception as e:
            session.mark_failed(str(e))
            raise
    
    @staticmethod
    def get_upload_templates() -> Dict[str, UploadTemplate]:
        """Get all available upload templates"""
        templates = {}
        for template in UploadTemplate.objects.filter(is_active=True):
            templates[template.upload_type] = template
        return templates
    
    @staticmethod
    def validate_file_headers(upload_type: str, headers: List[str]) -> Tuple[bool, List[str]]:
        """Validate file headers against template requirements"""
        try:
            template = UploadTemplate.objects.get(upload_type=upload_type, is_active=True)
            missing_headers = template.validate_headers(headers)
            return len(missing_headers) == 0, missing_headers
        except UploadTemplate.DoesNotExist:
            return True, []  # No template validation